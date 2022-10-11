#%%
from datetime import timedelta
from email.mime import image
from time import time

from tkinter import *
from turtle import left
from PIL import ImageTk, Image

import xlwings as xw
import pandas as pd
from openpyxl import workbook
from  openpyxl import load_workbook
from tkinter import messagebox
from tkinter import *
from tkinter.ttk import *
from tkinter.filedialog import askopenfile

# import openpyxl
# import datetime
from datetime import date
import win32com.client as win32
import os


root = Tk()
root.geometry('750x300')
root.resizable(False,False)
root.title('LaraExtr_Reminder')



def open_file():
    global RwFilePath,mis_tbl,vyg
    
    file = askopenfile(mode='r',filetypes=[('Excel Files','*.XLSX')])
    RwFilePath = file.name
    pthLbl.configure(text = RwFilePath, font= ('Helvetica 10 italic'))
    # messagebox.showinfo(RwFilePath,"Excel Tracker path")

def Process_tbl():

    df = pd.read_excel(RwFilePath)
    lst = df[df.columns[0]].to_list()
    rwCnt = (lst.index('Line')+1)

    dfTrack = pd.read_excel(RwFilePath,skiprows=rwCnt)

    df_missing = dfTrack[dfTrack['Status']=='MISSING']

    # df_voy = df_missing['VOYAGE','PORT'].drop_duplicates()
    df_voy = df_missing.drop_duplicates(subset = ['VOYAGE','PORT'],keep = 'last').reset_index(drop = True)

    max_port = df_voy[df_voy['PORT'].str.len()>5]
    max_port.to_excel("Multiple_calls.xlsx")
    # omit and na
    xmacro = xw.Book('LC_ETD_reminder_Python.xlsm')
    macroRun = xmacro.macro('Module1.Clear_Sheet')
    macroRun()

    
    sheet = xmacro.sheets['HOME']
    lstVyg = df_voy['VOYAGE'].to_list()
    strVyg = "','".join([elm for elm in lstVyg])
    sheet.range('A2').value = strVyg

    macroRun = xmacro.macro('Module3.SQL_Query_omit_NA')
    macroRun()
    xmacro.save()
    xmacro.close()


    path1 = './LC_ETD_reminder_Python.xlsm'
    sheet_name = 'RESULT'
    Rslt = load_workbook(path1, keep_vba=True)
    df_omit = pd.DataFrame(Rslt['RESULT'].values)
    df_omit.rename(columns=df_omit.iloc[0], inplace = True)

    df_omit.to_excel("Omit_Details.xlsx")

    df_omit = df_omit.drop_duplicates(subset=['VOYAGE_REFERENCE','POINT_CODE','PORT_ACTIVITY'],keep = 'last').reset_index(drop = True)
    df_omit['Uniqu_keys']= df_omit['VOYAGE_REFERENCE']+ df_omit['POINT_CODE']
    duplicate_val = df_omit[df_omit['Uniqu_keys'].duplicated()]
    uniq_val = df_omit[~(df_omit['Uniqu_keys'].isin(duplicate_val['Uniqu_keys'].to_list()))]
    duplicate_Cnt = df_omit[(df_omit['Uniqu_keys'].isin(duplicate_val['Uniqu_keys'].to_list()))]
    
    duplicate_final = duplicate_Cnt[(duplicate_Cnt['PORT_ACTIVITY'] == "O") | (duplicate_Cnt['PORT_ACTIVITY'] == "D")]
    # rslt = uniq_val.append(duplicate_final)
    rslt = uniq_val[(uniq_val['PORT_ACTIVITY'] == "O") + (uniq_val['PORT_ACTIVITY'] == "D")]


    df_omit_TBN = df_omit[df_omit['FULL_NAME'].str.contains('TBN')]



    # df_missing = dfTrack[dfTrack['Status']=='MISSING']
    # df_missing['VOYAGE']
    # df_missing = pd.merge(df_missing,df_omit_vyg, left_on=['VOYAGE','PORT'],right_on=['VOYAGE_REFERENCE','POINT_CODE'],how='left')

    df_missing['Unique_key'] = df_missing['VOYAGE'] +"|"+df_missing['PORT']
    rslt['Unique_key'] = rslt['VOYAGE_REFERENCE'] +'|'+rslt['POINT_CODE']
    df_omit_vyg_List = rslt['Unique_key'].drop_duplicates(keep='last').to_list()

    df_omit_TBN['Unique_key'] = df_omit_TBN['VOYAGE_REFERENCE'] +'|'+df_omit_TBN['POINT_CODE']
    df_omit_TBN_List = df_omit_TBN['Unique_key'].drop_duplicates(keep='last').to_list()

    pattern = '#'.join(df_omit_vyg_List)
    pattern1 = '#'.join(df_omit_TBN_List)


######################################

    if len(pattern) >0:
        # duplicate_final
        df_missing = df_missing[~(df_missing['Unique_key'].isin(duplicate_final['Uniqu_keys']))]        
        df_missing = df_missing[~(df_missing['Unique_key'].isin(rslt['Unique_key']))]        
    if len(pattern1) >0:
        df_missing = df_missing[~(df_missing['Unique_key'].isin(df_omit_TBN['Unique_key']))]        
        # df_missing = df_missing[~df_missing['Unique_key'].str.contains(pattern1,case=False)]

    df_voy = df_missing['VOYAGE'].drop_duplicates()


    # next

    xmacro = xw.Book('LC_ETD_reminder_Python.xlsm')
    macroRun = xmacro.macro('Module1.Clear_Sheet')    
    macroRun()


    sheet = xmacro.sheets['HOME']
    lstVyg = df_voy.to_list()
    strVyg = "','".join([elm for elm in lstVyg])
    sheet.range('A2').value = strVyg

    macroRun = xmacro.macro('Module1.SQL_Query')
    macroRun()
    xmacro.save()
    xmacro.close()


    path1 = './LC_ETD_reminder_Python.xlsm'
    sheet_name = 'RESULT'
    Rslt = load_workbook(path1, keep_vba=True)
    df_status = pd.DataFrame(Rslt['RESULT'].values)
    df_status.rename(columns=df_status.iloc[0], inplace = True)

###################################### GMT to local###############3
    df_status.to_excel('OUTPUT.xlsx')

    rst_xl = pd.read_excel("OUTPUT.xlsx")

    for ind,rw in rst_xl.iterrows():
        lngth = len(str(rw['ETD_TIME']))
        if lngth == 4:
            rst_xl.loc[ind,'Hrs'] = str(rw['ETD_TIME'])[0:2]
            rst_xl.loc[ind,'Mnts'] = str(rw['ETD_TIME'])[-2:]

            addHrs1 = rst_xl.loc[ind,'ETD_DATE'] + timedelta(hours=int(str(rw['ETD_TIME'])[0:2]),minutes=int(str(rw['ETD_TIME'])[-2:])) 
            rst_xl.loc[ind,'ETD_DATE'] = addHrs1 +timedelta(hours=5,minutes=30,seconds=0)

        elif lngth == 3:
            rst_xl.loc[ind,'Hrs'] = str(rw['ETD_TIME'])[0:1]
            rst_xl.loc[ind,'Mnts'] = str(rw['ETD_TIME'])[-2:]
            addHrs = rst_xl.loc[ind,'ETD_DATE'] + timedelta(hours=int(str(rw['ETD_TIME'])[0:1]),minutes=int(str(rw['ETD_TIME'])[-2:])) 
            rst_xl.loc[ind,'ETD_DATE'] = addHrs + +timedelta(hours=5,minutes=30,seconds=0)       #GMT+5.30
    rst_xl.to_excel('OUTPUT.xlsx')

    df_status = pd.read_excel("OUTPUT.xlsx")

######################################




    # delete 1st row
    df_status.drop(index=df_status.index[0],axis=0,inplace=True)
    dfNew = pd.to_datetime(df_status['ETD_DATE']).dt.date           #convert to date time format

    df_status['rev__']=dfNew
    yest_Details_LARA = df_status[df_status['rev__'] < date.today()]


####ach
    yest_Details_LARA['UNK'] = yest_Details_LARA['VOYAGE_REFERENCE'] +"|"+ yest_Details_LARA['POINT_CODE']
    yest_Details_LARA['a_type_date'] = [max(yest_Details_LARA[(yest_Details_LARA['UNK']==i)]['rev__']) for i in yest_Details_LARA['UNK']]
    yest_Details_LARA['TRU_STATUS'] = ""
    yest_Details_LARA.loc[yest_Details_LARA['a_type_date'] == yest_Details_LARA['rev__'],'TRU_STATUS'] = True
    yest_Details_LARA = yest_Details_LARA[yest_Details_LARA['TRU_STATUS'] == True]

    yest_Details_LARA = yest_Details_LARA.drop_duplicates(subset = ['VOYAGE_REFERENCE','POINT_CODE'],keep = 'last').reset_index(drop = True)

    future_Details_Fut = df_status[df_status['rev__'] >= date.today()]
    future_Details_Fut.to_excel("Future_Details.xlsx")

    

    # df_status.rename(columns=df_status.iloc[0], inplace = True)
    # df_status['Unique_key'] = df_status['VOYAGE_REFERENCE'] +"|"+df_status['POINT_CODE']
    # dfTrack['Unique_key'] = dfTrack['VOYAGE'] +'|'+dfTrack['PORT']

    # Tracker file 
    rev_missing = df_missing[['VOYAGE','VESSEL NAME','Vessel Code','PORT','Service']]
    rev_missing = rev_missing.sort_values(['VOYAGE','PORT']).drop_duplicates('VOYAGE',keep='last')   
    # rev_missing.drop_duplicates(keep=False,inplace=True)
    # Left_Join = pd.merge(dfTrack,df_status,on='Unique_key',how='left')
    Left_Join = pd.merge(rev_missing,yest_Details_LARA,left_on=['VOYAGE','PORT'],right_on=['VOYAGE_REFERENCE','POINT_CODE'],how='inner')
    

    Left_Join['Days_diff'] = ''
    for index,row in Left_Join.iterrows():
        # print(date.today())
        # print(row['ETD_DATE']) 
        if pd.isna(row['ETD_DATE']) == False:          
            LaraDate = row['ETD_DATE']
            dayDiff = LaraDate.date() - date.today()
            # print(dayDiff.days)
            Left_Join.loc[index,'Days_diff'] = dayDiff.days
    
    mdf = pd.read_excel('LC_Email.xlsx',sheet_name='Contact_Ids')
    Left_Join = pd.merge(Left_Join,mdf,how='left',on='PORT')
    Left_Join['From'] = "SSC.bayplan@cma-cgm.com"
    Left_Join['CC'] = "SSC.RAPOOJARY@cma-cgm.com;"

    Left_Join.to_excel("OUTPUT.xlsx")
    messagebox.showinfo("Done!!","Thank you")

def Edi_data():
    dfTrack =  pd.read_excel('OUTPUT.xlsx')

    # df_missing = dfTrack[dfTrack['Status']=='MISSING']

    df_voy = dfTrack['VOYAGE'].drop_duplicates()

    xmacro = xw.Book('LC_ETD_reminder_Python.xlsm')
    macroRun = xmacro.macro('Module1.Clear_Sheet')
    macroRun()


    sheet = xmacro.sheets['HOME']
    lstVyg = df_voy.to_list()
    strVyg = "','".join([elm for elm in lstVyg])
    sheet.range('A2').value = strVyg

    macroRun = xmacro.macro('Module2.SQL_Query_edi')
    macroRun()
    xmacro.save()
    xmacro.close()

    path1 = './LC_ETD_reminder_Python.xlsm'
    sheet_name = 'RESULT'
    Rslt = load_workbook(path1, keep_vba=True)
    df_status = pd.DataFrame(Rslt['RESULT'].values)
    df_status.rename(columns=df_status.iloc[0], inplace = True)
    df_status.drop(index=df_status.index[0],axis=0,inplace=True)
    # df_status.rename(columns=df_status.iloc[0], inplace = True)
    # df_status['Unique_key'] = df_status['VOYAGE_REFER__EDI']
    # dfTrack['Unique_key'] = dfTrack['VOYAGE']

    Left_Join = pd.merge(dfTrack,df_status,left_on=['VOYAGE'],right_on=['VOYAGE_REFER__EDI'],how='left')
    

    edi_Email = pd.read_excel('EDI Emails.xlsx')
    edi_Email.columns = edi_Email.columns.astype("str")
    # edi_Email.columns = edi_Email.columns.str.title()

    getEm = pd.merge(Left_Join,edi_Email,left_on="Service",right_on='SERVICE',how='left')
    # getEm['PARTNER_NAME_EDI'].dropna(inplace = True)
    getEm['PARTNER_NAME_EDI'].fillna("##",inplace=True)
    getEm['OPERATOR'].fillna("#12#",inplace=True)

    for idx,rw in getEm.iterrows():
        if getEm.loc[idx,'PARTNER_NAME_EDI'].find(getEm.loc[idx,'OPERATOR']) >-1:        
            getEm.loc[idx,'IS_Exist'] = 'Yes'
        else:
            getEm.loc[idx,'IS_Exist'] = 'No'

    ndf= getEm[getEm['IS_Exist']=='Yes']
    ndf_NOT= getEm[getEm['IS_Exist']=='No']
    ndf.to_excel("OUTPUT.xlsx")
    ndf_NOT.to_excel("Email_not_found.xlsx")
    messagebox.showinfo("Done!!","Thank you")

def EDI_mail():
    
    Left_Join = pd.read_excel('OUTPUT.xlsx')
    outlook = win32.Dispatch('outlook.application')
    oacctouse = None
    for index,row in Left_Join.iterrows():
        if int(Left_Join.loc[index,'Days_diff']) <=-1:    
            frmeml = Left_Join.loc[index,'From']
            CC = Left_Join.loc[index,'CC']
            # eml = 'ssc.achauhan@cma-cgm.com'
            eml = Left_Join.loc[index,'E-MAIL ID']
            for oacc in outlook.Session.Accounts:
                if oacc.SmtpAddress == frmeml:
                    oacctouse = oacc
                    break

            outlookNm = outlook.GetNameSpace("MAPI")
            mail = outlook.CreateItem(0)
            # mail.SendUsingAccount = oacctouse
            if oacctouse:
                mail._oleobj_.Invoke(*(64209, 0, 8, 0, oacctouse))

            # mail.From = frmeml
            mail.To = eml
            mail.CC = CC
            mail.Subject = "~Reminder 1: DEPARTURE PARTNER EDI FILE REQUIRED FOR SERVICE - # " + Left_Join.loc[index,'SERVICE'] + " " +Left_Join.loc[index,'VESSEL NAME'] +" "+ Left_Join.loc[index,'VOYAGE']+" "+ Left_Join.loc[index,'PORT']

            strVal = "Can you please check and send departure EDI of subject vessel/voyage/Port. "
            strVal2 = "Also please advise if there is no load no discharge."
            strVal3 = "Note : If already sent, Kindly ignore the mails.."
            mail.HTMLBody = "<p>Good Day Team,</p></br>{}<br><br>{}<br><br>{}</br><p>Thank you.</p><br>Regards,<br>GBS Mumbai<br>CMA-CGM GLOBAL BUSINESS SERVICES INDIA<br>".format(strVal,strVal2,strVal3)

            Left_Join.loc[index,'Remark'] = "Sent"
            Left_Join.loc[index,'Subject_Line'] = mail.Subject

            # mail.display()
            mail.Send()        
    
    Left_Join.to_excel("OUTPUT.xlsx")
    messagebox.showinfo("Done!!","Thank you")

def Non_EDI_Mail_dist():

    Left_Join = pd.read_excel('OUTPUT.xlsx')
    outlook = win32.Dispatch('outlook.application')
    oacctouse = None
    for index,row in Left_Join.iterrows():
        # if Left_Join.loc[index,'Status'] == 'MISSING' and int(Left_Join.loc[index,'Days_diff']) <=-1:    
        if int(Left_Join.loc[index,'Days_diff']) <=-1:    
            frmeml = Left_Join.loc[index,'From']
            CC = Left_Join.loc[index,'CC']
            # eml = 'ssc.achauhan@cma-cgm.com'
            eml = Left_Join.loc[index,'emails']
            for oacc in outlook.Session.Accounts:
                if oacc.SmtpAddress == frmeml:
                    oacctouse = oacc
                    break

            outlookNm = outlook.GetNameSpace("MAPI")
            mail = outlook.CreateItem(0)
            # mail.SendUsingAccount = oacctouse
            if oacctouse:
                mail._oleobj_.Invoke(*(64209, 0, 8, 0, oacctouse))

            # mail.From = frmeml
            mail.To = eml
            mail.CC = CC
            mail.Subject = "~Reminder 1: Load Confirmation required # " + Left_Join.loc[index,'VESSEL NAME'] +" "+ Left_Join.loc[index,'VOYAGE']+" "+ Left_Join.loc[index,'PORT']

            strVal = "Can you please urgently assist to provide Loading Confirmation for subject vessel including CMA Group Boxes(CXML format) in order to integrate bayplan."
            strVal2 = "Note : If already sent, Kindly ignore the mails.."
            mail.HTMLBody = "<p>Good Day Team,</p></br>{}<br><br><br>{}</br><p>Thank you.</p><br>Regards,<br>GBS Mumbai<br>CMA-CGM GLOBAL BUSINESS SERVICES INDIA<br>".format(strVal,strVal2)

            Left_Join.loc[index,'Remark'] = "Sent"
            Left_Join.loc[index,'Subject_Line'] = mail.Subject

            # mail.display()
            mail.Send()        
    
    Left_Join.to_excel("OUTPUT.xlsx")
    messagebox.showinfo("Done!!","Thank you")

lgnLbl = Label(root,text="Welcome :" + os.getlogin())
lgnLbl.place(x=550,y=5)

pthLbl = Label(root,text='Choose File')
pthLbl.place(x=0,y=20)


canvas = Canvas(width=250, height=210, bg='blue')


canvas.pack(expand=YES, fill=X)


image = ImageTk.PhotoImage(file="VSL.jpg")
canvas.create_image(15, 5, image=image, anchor=NW)

btn = Button(root, text ='Select Tracker File', command = lambda:open_file())
btn.place(x=630, y=60)


btn_Proc = Button(root, text ='LC_Tracker_Process', command = lambda:Process_tbl())
btn_Proc.place(x=630, y=100)


btn_Proc_EDI = Button(root, text ='EDI_Process', command = lambda:Edi_data())
btn_Proc_EDI.place(x=630, y=140)


btn_Mail = Button(root, text ='LC_Mail_Dist', command = lambda:Non_EDI_Mail_dist())
btn_Mail.place(x=630, y=180)


btn_Mail_e = Button(root, text ='EDI_Mail_Dist', command = lambda:EDI_mail())
btn_Mail_e.place(x=630, y=220)



mainloop()


#%%

