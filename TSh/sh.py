
#%%
from re import T
import pandas as pd
import tkinter as tk
from tkinter import N, filedialog
import os
from datetime import datetime
from tkinter import messagebox
from tkinter import *
from tkinter.ttk import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Create Object
root = Tk()
 
# Set geometry (widthxheight)
root.geometry('500x200')
 
# This will create style object
style = Style()

style.configure('W.TButton', font =
               ('calibri', 10, 'bold', 'underline'),
                foreground = 'red')



def browseFoldr(ttl):
    fldr = filedialog.askdirectory(title=ttl)
    return fldr 


def background_colors(path):
    wb = load_workbook('Result.xlsx') 
    sheet = wb.active

    rw_range = sheet.max_row
    col_range = sheet.max_column

    for cl in range(2,(col_range+1)):
        for rw in range(4,(rw_range+1)):
            if sheet.cell(rw,cl).value <0:
                sheet.cell(rw,cl).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
            else:
                sheet.cell(rw,cl).fill = PatternFill(start_color='00B050', end_color='00B050', fill_type="solid")

    wb.save(path)


def process():
    pth = browseFoldr('StronglyOutperforming Files')
    pthNSE = browseFoldr('NSE Files.')

    dfCom = pd.DataFrame(columns=['Stock Name','Symbol','Exch','Sector Name','Industry Name','Index','RS','Close','Chg %','Symbol with Comma for External Upload','Date_Based'
    ])
    # dfCom= pd.DataFrame()
    for file in os.listdir(pth):
        if file.endswith('.csv'):
            f = os.path.join(pth,file)
            df = pd.read_csv(f,sep=';')
            splVal = (df.iloc[0].str.split(":",regex=True))
            dt = splVal.iloc[0][1].replace(',','')
            dt1 = dt.strip()
            dfN = pd.read_csv(f,sep=',',skiprows=4)
            dfN['Date_Based'] = dt1
            # dfCom = pd.merge(dfCom,dfN,left_on='Stock Name',right_on='Stock Name',how='outer',indicator=True)
            # dfN.to_excel(os.path.join("Rev_"+ file.replace('.csv','.xlsx')))
            dfCom = pd.concat([dfCom,dfN])

    # pvTbl = dfCom.pivot(index='Stock Name',columns='Date_Based',values=['RS','Close','Chg %'])

    dfComNSE = pd.DataFrame(columns=['Symbol","Series","Date","Prev Close","Open Price","High Price","Low Price","Last Price","Close Price","Average Price","Total Traded Quantity","Turnover","No. of Trades","Deliverable Qty","% Dly Qt to Traded Qty'])

    # dfCom= pd.DataFrame()
    for file in os.listdir(pthNSE):
        if file.endswith('.csv'):
            f = os.path.join(pthNSE,file)
            df = pd.read_csv(f,sep=',')
            # final_df = df.sort_values(by='Date', ascending=False)
            final_df = df
            if len(final_df[final_df['Deliverable Qty'] == '-']) >0:        
                final_df = final_df.replace('-','0')
                final_df['Deliverable Qty'] = final_df['Deliverable Qty'].str.strip()        
                final_df['Deliverable Qty'] = pd.to_numeric(final_df['Deliverable Qty'])

            if len(final_df[final_df['Total Traded Quantity'] == '-']) >0:        
                final_df = final_df.replace('-','0')
                final_df['Total Traded Quantity'] = final_df['Total Traded Quantity'].str.strip()        
                final_df['Total Traded Quantity'] = pd.to_numeric(final_df['Total Traded Quantity'])

            final_df['Date'] = pd.to_datetime(final_df['Date'])
            final_df.sort_values(by='Date', ascending=True,inplace=True)
            

            final_df['Trd_Diff'] = final_df['Total Traded Quantity'].diff()
            final_df['DelQ_Diff'] = final_df['Deliverable Qty'].diff()

            dfComNSE = pd.concat([dfComNSE,final_df])

    dfComNSE['Date'] = pd.to_datetime(dfComNSE['Date'], format='%Y-%m-%d')
    dfComNSE['Date'] = dfComNSE['Date'].dt.strftime('%d-%b-%Y')

    fTbl = pd.merge(left=dfCom,right=dfComNSE,left_on=['Symbol','Date_Based'],right_on=['Symbol','Date'],how='outer',indicator=True)
    fTbl = fTbl[(fTbl['_merge']=='both')|(fTbl['_merge']=='right_only')]
    fTbl['Date']=pd.to_datetime(fTbl['Date'])

    pvTbl = pd.pivot_table(fTbl,index=['Symbol'],columns='Date',values=['RS','Close','Chg %','Trd_Diff','DelQ_Diff'])

    for i in pvTbl.columns:
        pvTbl.rename(columns={i[1]:i[1].strftime('%d-%b-%Y')},inplace=True)

    pvTbl.to_excel('Result.xlsx')

    background_colors("Result.xlsx")
    messagebox.showinfo('Thank you','Completed..')
# Style will be reflected only on
# this button because we are providing
# style only on this Button.
''' Button 1'''
btn1 = Button(root, text = 'Quit !',
                style = 'W.TButton',
             command = root.destroy)
btn1.grid(row = 0, column = 3, padx = 100)
 
''' Button 2'''
 
btn2 = Button(root, text = 'Click me !', command = process)
btn2.grid(row = 1, column = 3, pady = 10, padx = 100)
 
# Execute Tkinter
root.mainloop()


#%%




# # Create Object
# root = Tk()
 
# # Set geometry (widthxheight)
# root.geometry('500x200')
 
# # This will create style object
# style = Style()

# style.configure('W.TButton', font =
#                ('calibri', 10, 'bold', 'underline'),
#                 foreground = 'red')

# def browseFoldr(ttl):
#     fldr = filedialog.askdirectory(title=ttl)
#     return fldr 
    

# def process():
#     pth = browseFoldr('StronglyOutperforming Files')
#     pthNSE = browseFoldr('NSE Files.')
    
#     dfCom = pd.DataFrame(columns=['Stock Name','Symbol','Exch','Sector Name','Industry Name','Index','RS','Close','Chg %','Symbol with Comma for External Upload','Date_Based'
#     ])
#     # dfCom= pd.DataFrame()
#     for file in os.listdir(pth):
#         if file.endswith('.csv'):
#             f = os.path.join(pth,file)
#             df = pd.read_csv(f,sep=';')
#             splVal = (df.iloc[0].str.split(":",regex=True))
#             dt = splVal.iloc[0][1].replace(',','')
#             dt1 = dt.strip()
#             dfN = pd.read_csv(f,sep=',',skiprows=4)
#             dfN['Date_Based'] = dt1
#             # dfCom = pd.merge(dfCom,dfN,left_on='Stock Name',right_on='Stock Name',how='outer',indicator=True)
#             # dfN.to_excel(os.path.join("Rev_"+ file.replace('.csv','.xlsx')))
#             dfCom = pd.concat([dfCom,dfN])

#     # pvTbl = dfCom.pivot(index='Stock Name',columns='Date_Based',values=['RS','Close','Chg %'])

#     dfComNSE = pd.DataFrame(columns=['Symbol","Series","Date","Prev Close","Open Price","High Price","Low Price","Last Price","Close Price","Average Price","Total Traded Quantity","Turnover","No. of Trades","Deliverable Qty","% Dly Qt to Traded Qty'])

#     # dfCom= pd.DataFrame()
#     for file in os.listdir(pthNSE):
#         if file.endswith('.csv'):
#             f = os.path.join(pthNSE,file)
#             df = pd.read_csv(f,sep=',')
#             # final_df = df.sort_values(by='Date', ascending=False)
#             final_df = df
#             if len(final_df[final_df['Deliverable Qty'] == '-']) >0:        
#                 final_df = final_df.replace('-','0')
#                 final_df['Deliverable Qty'] = final_df['Deliverable Qty'].str.strip()        
#                 final_df['Deliverable Qty'] = pd.to_numeric(final_df['Deliverable Qty'])

#             if len(final_df[final_df['Total Traded Quantity'] == '-']) >0:        
#                 final_df = final_df.replace('-','0')
#                 final_df['Total Traded Quantity'] = final_df['Total Traded Quantity'].str.strip()        
#                 final_df['Total Traded Quantity'] = pd.to_numeric(final_df['Total Traded Quantity'])

#             final_df['Date'] = pd.to_datetime(final_df['Date'])
#             final_df.sort_values(by='Date', ascending=True,inplace=True)
            

#             final_df['Trd_Diff'] = final_df['Total Traded Quantity'].diff()
#             final_df['DelQ_Diff'] = final_df['Deliverable Qty'].diff()

#             dfComNSE = pd.concat([dfComNSE,final_df])

#     dfComNSE['Date'] = pd.to_datetime(df.Date, format='%d-%m-%Y')
#     dfComNSE['Date'] = dfComNSE['Date'].dt.strftime('%d-%b-%Y')

#     fTbl = pd.merge(left=dfCom,right=dfComNSE,left_on=['Symbol','Date_Based'],right_on=['Symbol','Date'],how='inner',indicator=True)
#     fTbl = fTbl[fTbl['_merge']=='both']

#     # # dfComNSE['Date'] = dfComNSE['Date'].dt.strftime('%d-%b-%Y')
#     pvTbl = pd.pivot_table(fTbl,index=['Stock Name','Symbol'],columns='Date_Based',values=['RS','Close','Chg %','Trd_Diff','DelQ_Diff'])
#     pvTbl.to_excel('Result.xlsx')

#     print('dd')

# # Style will be reflected only on
# # this button because we are providing
# # style only on this Button.
# ''' Button 1'''
# btn1 = Button(root, text = 'Quit !',
#                 style = 'W.TButton',
#              command = root.destroy)
# btn1.grid(row = 0, column = 3, padx = 100)
 
# ''' Button 2'''
 
# btn2 = Button(root, text = 'Click me !', command = process)
# btn2.grid(row = 1, column = 3, pady = 10, padx = 100)
 
# # Execute Tkinter
# root.mainloop()



